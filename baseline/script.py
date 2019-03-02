import django, os, sys
sys.path.append('E:\demo\\python\\beilin_phase2\\baseline')
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "beilin_phase2.settings")
django.setup()


from baseline.models import Companies
# from accounts.models import User
# from django.contrib.auth.models import Group
from django.utils import timezone
from openpyxl import load_workbook

if __name__ == '__main__':

    companies = Companies.objects.filter(complete_done=True, is_valid=True)
    for company in companies:
        if not company.submit_date and company.modified_date:
            Companies.objects.filter(id=company.id).update(submit_date=company.modified_date)

    print('11111')


'''
    students = User.objects.all().exclude(is_admin=True)
    g = Group.objects.get(name='考试')

    for student in students:
        student.district = '南院门'
        g.user_set.add(student)
        student.save()

'''




'''
    base_folder = "E:\demo\\python\\beilin_phase2\\baseline\\"
    file_name = "经济普查考核人员名单.xlsx"
    wb = load_workbook(base_folder+file_name)

    ws = wb.active
    for row in ws.iter_rows(min_row=3, max_row=155):
        name = row[1].value
        mobile = str(row[3].value)
        available_date = row[2].value
        password = "TXHQbljp" + str(mobile).strip()[7:11]
        print(name, mobile, password)

        # user = User.objects.create_user(name=name, mobile=mobile, password=password, district='')
        # user.is_staff = True
        user = User.objects.get(mobile=mobile)
        user.available_date = available_date
        user.save()


    street_district_ls = ['南院门', '柏树林']

    for sd in street_district_ls:
        print(sd)

        cc = Companies.objects.filter(street_district__contains=sd)
        if cc:
            for c in cc:
                c.street_district = sd
                c.save()
'''